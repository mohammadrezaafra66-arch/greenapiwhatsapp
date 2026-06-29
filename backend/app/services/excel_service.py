"""
Excel import/export service for contacts.
"""
import io
import re
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill

def normalize_phone(phone: str) -> Optional[str]:
    """Normalize Iranian phone numbers to 989xxxxxxxxx format."""
    if not phone:
        return None
    phone = str(phone).strip().replace("+", "").replace("-", "").replace(" ", "")
    # Remove any non-digit characters
    phone = re.sub(r"\D", "", phone)
    if not phone:
        return None
    # Convert formats
    if phone.startswith("0") and len(phone) == 11:
        phone = "98" + phone[1:]
    elif len(phone) == 10 and phone.startswith("9"):
        phone = "98" + phone
    elif phone.startswith("98") and len(phone) == 12:
        pass  # Already correct
    else:
        return None  # Invalid format
    # Validate Iranian mobile
    if not re.match(r"^989[0-9]{9}$", phone):
        return None
    return phone

def parse_contacts_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse Excel file with contacts.
    Expected columns: phone, first_name, last_name, province, city
    Columns can be in any order if headers are correct.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[str(cell_value).lower().strip()] = col

    # Map possible column names
    phone_col = headers.get("phone") or headers.get("شماره") or headers.get("موبایل") or 1
    fname_col = headers.get("first_name") or headers.get("نام") or headers.get("اسم")
    lname_col = headers.get("last_name") or headers.get("فامیلی") or headers.get("نام خانوادگی")
    province_col = headers.get("province") or headers.get("استان")
    city_col = headers.get("city") or headers.get("شهر")

    contacts = []
    seen_phones = set()

    for row in range(2, ws.max_row + 1):
        raw_phone = ws.cell(row, phone_col).value
        if not raw_phone:
            continue

        phone = normalize_phone(str(raw_phone))
        if not phone or phone in seen_phones:
            continue
        seen_phones.add(phone)

        contact = {"phone": phone}
        if fname_col:
            contact["first_name"] = ws.cell(row, fname_col).value
        if lname_col:
            contact["last_name"] = ws.cell(row, lname_col).value
        if province_col:
            contact["province"] = ws.cell(row, province_col).value
        if city_col:
            contact["city"] = ws.cell(row, city_col).value

        contacts.append(contact)

    return contacts


def export_logs_excel(logs: list[dict]) -> bytes:
    """Export send logs as Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Send Logs"

    headers = ["شماره", "نام", "وضعیت", "زمان ارسال", "حساب", "Message ID", "خطا"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row_idx, 1, log.get("phone", ""))
        ws.cell(row_idx, 2, log.get("name", ""))
        ws.cell(row_idx, 3, log.get("status", ""))
        ws.cell(row_idx, 4, str(log.get("sent_at", "")))
        ws.cell(row_idx, 5, log.get("account_name", ""))
        ws.cell(row_idx, 6, log.get("message_id", ""))
        ws.cell(row_idx, 7, log.get("error", ""))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
